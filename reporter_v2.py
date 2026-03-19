"""
reporter_v2.py
==============
VERSION 2 — Professional Excel Workpaper Generator
Big 4-standard multi-sheet workpaper with Going Concern and Benford's Law sheets

Author: Purva Doshi's Audit AP Tool v2.0
"""

import pandas as pd
from datetime import date
import io

# Lazy-load openpyxl only when generate_report is called (not at import time)
# This prevents crashes on Python 3.14 where openpyxl may not be available at startup
_openpyxl_loaded = False

def _ensure_openpyxl():
    global _openpyxl_loaded, Workbook, Font, PatternFill, Alignment, Border, Side, get_column_letter
    if not _openpyxl_loaded:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        _openpyxl_loaded = True


# ── Palette ──────────────────────────────────────────────────
P = {
    "navy":        "0F2B4C",   # Deep navy header
    "navy_mid":    "1A3F6F",   # Mid navy subheader
    "slate":       "334155",
    "red_bg":      "FEE2E2",   # Risk row backgrounds
    "yellow_bg":   "FEF9C3",
    "green_bg":    "DCFCE7",
    "grey_alt":    "F8FAFC",
    "white":       "FFFFFF",
    "divider":     "E2E8F0",
    "gold":        "F59E0B",   # Accent
    "text_dark":   "0F172A",
    "text_mid":    "475569",
    "text_light":  "94A3B8",
    "gc_critical": "7F1D1D",   # Going concern critical
    "gc_high":     "991B1B",
    "gc_mod":      "92400E",
    "gc_low":      "14532D",
}

def _thin_border(color="CBD5E1"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    return Border(bottom=Side(style="medium", color="94A3B8"))

BORDER = _thin_border()


def _h(ws, row, col_count, text, fg=P["navy"], font_color="FFFFFF",
        size=11, bold=True, height=28, indent=0):
    """Write a full-width merged header row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    c = ws.cell(row=row, column=1, value=text)
    c.fill  = PatternFill("solid", fgColor=fg)
    c.font  = Font(name="Calibri", size=size, bold=bold, color=font_color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent + 1)
    c.border = BORDER
    ws.row_dimensions[row].height = height
    return c


def _col_widths(ws, widths: dict):
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _row_bg(ws, row_num, col_count, bg_color, height=18):
    ws.row_dimensions[row_num].height = height
    for col in range(1, col_count + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.border = BORDER


def _flag_bg(flag: str) -> str:
    if "🔴" in str(flag):  return P["red_bg"]
    if "🟡" in str(flag):  return P["yellow_bg"]
    if "🟢" in str(flag):  return P["green_bg"]
    return P["white"]


# ══════════════════════════════════════════════════════════════
# SHEET 1 — COVER & SUMMARY
# ══════════════════════════════════════════════════════════════

def _cover_sheet(ws, summary: dict, client_name: str, period: str, unit_label: str, standard: str):
    ws.title = "1. Cover"
    ws.sheet_view.showGridLines = False
    N = 8

    # ── Title block ──
    ws.row_dimensions[1].height = 48
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "AUDIT ANALYTICAL REVIEW"
    c.font  = Font(name="Calibri", size=22, bold=True, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=P["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = f"{summary.get('Standard', standard)} — Analytical Procedures Workpaper | Version 2.0"
    c2.font  = Font(name="Calibri", size=10, italic=True, color=P["text_light"])
    c2.fill  = PatternFill("solid", fgColor=P["navy_mid"])
    c2.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 8

    # ── Engagement details ──
    _h(ws, 4, N, "ENGAGEMENT DETAILS", fg=P["navy_mid"], size=10, height=20)
    details = [
        ("Client:",         client_name),
        ("Period / Year End:", period),
        ("Reporting Standard:", summary.get("Standard", standard)),
        ("Amounts in:",     unit_label),
        ("Report Date:",    date.today().strftime("%d %B %Y")),
        ("Prepared by:",    "Audit AP Tool v2.0"),
        ("File Reference:", "AP-WP-001 — Planning & Substantive"),
    ]
    for i, (lbl, val) in enumerate(details, start=5):
        ws.row_dimensions[i].height = 18
        lc = ws.cell(row=i, column=1, value=lbl)
        lc.font = Font(name="Calibri", size=10, bold=True, color=P["text_dark"])
        lc.fill = PatternFill("solid", fgColor=P["divider"])
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=N)
        vc = ws.cell(row=i, column=2, value=val)
        vc.font = Font(name="Calibri", size=10, color=P["text_dark"])
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border = BORDER

    ws.row_dimensions[13].height = 10

    # ── Key Metrics ──
    _h(ws, 14, N, "KEY FINANCIAL METRICS", fg=P["navy_mid"], size=10, height=20)

    def fmt(val, is_currency=True):
        if val is None: return "N/A"
        try:
            v = float(val)
            if is_currency:
                return f"{v:,.2f}"
            return f"{v:,.0f}"
        except Exception:
            return str(val)

    def rev_chg():
        cy, py = summary.get("CY Revenue", 0), summary.get("PY Revenue", 0)
        if py and py != 0:
            return f"{((cy-py)/abs(py)*100):+.1f}%"
        return "N/A"

    metrics = [
        ("Revenue (CY)",       fmt(summary.get("CY Revenue")),     "Revenue (PY)",       fmt(summary.get("PY Revenue"))),
        ("Net Profit (CY)",    fmt(summary.get("CY Net Profit")),   "Net Profit (PY)",    fmt(summary.get("PY Net Profit"))),
        ("Gross Profit (CY)",  fmt(summary.get("CY Gross Profit")), "Total Assets (CY)",  fmt(summary.get("CY Total Assets"))),
        ("Working Capital",    fmt(summary.get("CY Working Capital")), "Revenue Growth", rev_chg()),
        ("Total Accounts",     str(summary.get("Total Accounts in TB", 0)),
         "Unclassified A/c",   str(summary.get("Unclassified Accounts", 0))),
    ]

    # Header row
    ws.row_dimensions[15].height = 18
    for col, txt in enumerate(["Metric", "Value", "", "Metric", "Value", "", "", "Risk Overview"], start=1):
        c = ws.cell(row=15, column=col, value=txt)
        c.fill = PatternFill("solid", fgColor=P["slate"])
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for i, (lm, lv, rm, rv) in enumerate(metrics, start=16):
        ws.row_dimensions[i].height = 18
        bg = P["grey_alt"] if i % 2 == 0 else P["white"]
        for col, val in [(1, lm), (2, lv), (4, rm), (5, rv)]:
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10,
                           bold=(col in [1, 4]), color=P["text_dark"])
            c.alignment = Alignment(horizontal="left" if col in [1,4] else "right",
                                    vertical="center", indent=1)
            c.border = BORDER
        ws.cell(row=i, column=3).fill = PatternFill("solid", fgColor=P["divider"])
        ws.cell(row=i, column=3).border = BORDER

    # Risk overview box (right side)
    gc_risk = summary.get("Going Concern Risk", "LOW")
    flag_high = summary.get("Flagged — HIGH Risk", 0)
    flag_mod  = summary.get("Flagged — MODERATE", 0)

    gc_color = {"CRITICAL": "DC2626", "HIGH": "EA580C", "MODERATE": "D97706", "LOW": "16A34A"}.get(gc_risk, P["navy"])
    gc_icon  = {"CRITICAL": "🚨", "HIGH": "🔴", "MODERATE": "🟡", "LOW": "✅"}.get(gc_risk, "")

    ws.merge_cells("F16:H20")
    risk_cell = ws["F16"]
    risk_cell.value = f"{gc_icon}\nGoing Concern\n{gc_risk}\n\n🔴 HIGH: {flag_high}\n🟡 MOD: {flag_mod}"
    risk_cell.fill  = PatternFill("solid", fgColor=gc_color)
    risk_cell.font  = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    risk_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    risk_cell.border = BORDER

    ws.row_dimensions[22].height = 10
    _h(ws, 23, N, "SHEETS IN THIS WORKPAPER", fg=P["slate"], size=9, height=18)
    sheets_info = [
        ("1. Cover",          "This sheet — engagement details and summary"),
        ("2. Ratio Analysis", "15 financial ratios with risk flags and auditor notes"),
        ("3. Variance",       "All trial balance accounts — YoY movement analysis"),
        ("4. Flagged Items",  "Filtered view — HIGH and MODERATE risk accounts only"),
        ("5. Going Concern",  f"ISA/SA 570 assessment — {gc_risk} risk level"),
        ("6. Benford's Law",  "ISA/SA 240 fraud indicator — first-digit analysis (if uploaded)"),
        ("7. Instructions",   "Input format, categories, audit standards reference"),
    ]
    for i, (sh, desc) in enumerate(sheets_info, start=24):
        ws.row_dimensions[i].height = 16
        sc = ws.cell(row=i, column=1, value=sh)
        sc.font = Font(name="Calibri", size=9, bold=True, color=P["text_dark"])
        sc.fill = PatternFill("solid", fgColor=P["divider"])
        sc.border = BORDER
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=N)
        dc = ws.cell(row=i, column=2, value=desc)
        dc.font = Font(name="Calibri", size=9, color=P["text_mid"])
        dc.border = BORDER

    _col_widths(ws, {1: 26, 2: 18, 3: 3, 4: 26, 5: 18, 6: 3, 7: 3, 8: 20})


# ══════════════════════════════════════════════════════════════
# SHEET 2 — RATIO ANALYSIS
# ══════════════════════════════════════════════════════════════

def _ratio_sheet(ws, ratios: list, standard: str, unit_label: str):
    ws.title = "2. Ratio Analysis"
    ws.sheet_view.showGridLines = False
    N = 7
    headers = ["Ratio / Metric", "Formula", f"CY ({unit_label})", f"PY ({unit_label})",
               "YoY Change (%)", "Risk Flag", "Auditor's Note"]

    _h(ws, 1, N, f"RATIO ANALYSIS — {standard} | ISA/SA 520 Analytical Procedures",
       fg=P["navy"], size=12, height=32)

    # Column headers
    ws.row_dimensions[2].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=P["navy_mid"])
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    SECTIONS = {0: "📊 PROFITABILITY", 5: "💧 LIQUIDITY", 7: "⚖️ LEVERAGE & SOLVENCY", 9: "🔄 EFFICIENCY & STRUCTURE"}
    data_row = 3

    for idx, r in enumerate(ratios):
        if idx in SECTIONS:
            _h(ws, data_row, N, f"  {SECTIONS[idx]}", fg=P["divider"],
               font_color=P["text_dark"], size=9, bold=True, height=16)
            data_row += 1

        flag = str(r.get("Risk Flag", ""))
        bg = _flag_bg(flag)
        ws.row_dimensions[data_row].height = 20

        vals = [
            r.get("Ratio / Metric", ""),
            r.get("Formula", ""),
            r.get("Current Year"),
            r.get("Prior Year"),
            r.get("YoY Change (%)"),
            flag,
            r.get("Auditor's Note", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10, color=P["text_dark"],
                           bold=(col == 1))
            c.border = BORDER
            c.alignment = Alignment(
                horizontal="right" if col in [3, 4, 5] else "left",
                vertical="center", wrap_text=True, indent=0 if col in [3,4,5] else 1
            )
            if col in [3, 4]:
                name = r.get("Ratio / Metric", "")
                if isinstance(val, (int, float)) and val is not None:
                    if "%" in name or "Margin" in name or "ROE" in name or "Ratio" in name.split("Ratio")[0]:
                        c.number_format = "0.00"
                    elif "Working Capital" in name or "Revenue" in name:
                        c.number_format = "#,##0.00"
                    else:
                        c.number_format = "0.00"
            if col == 5 and isinstance(val, (int, float)):
                c.number_format = '+0.0;-0.0'
        data_row += 1

    # Legend
    data_row += 1
    _h(ws, data_row, N, "LEGEND", fg=P["slate"], size=9, height=16)
    data_row += 1
    for lbl, bg, desc in [
        ("🔴 HIGH — Investigate",  P["red_bg"],    "Material variance / significant risk → design specific audit procedures"),
        ("🟡 MODERATE — Explain",  P["yellow_bg"], "Unusual movement → obtain management explanation, document response"),
        ("🟢 OK",                  P["green_bg"],  "Within expected range → note and proceed"),
    ]:
        ws.row_dimensions[data_row].height = 16
        lc = ws.cell(row=data_row, column=1, value=lbl)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.font = Font(name="Calibri", size=9, bold=True)
        lc.border = BORDER
        ws.merge_cells(start_row=data_row, start_column=2, end_row=data_row, end_column=N)
        dc = ws.cell(row=data_row, column=2, value=desc)
        dc.font = Font(name="Calibri", size=9, color=P["text_mid"])
        dc.border = BORDER
        data_row += 1

    _col_widths(ws, {1: 32, 2: 42, 3: 16, 4: 16, 5: 13, 6: 28, 7: 52})
    ws.freeze_panes = ws.cell(row=3, column=1)


# ══════════════════════════════════════════════════════════════
# SHEET 3 & 4 — VARIANCE ANALYSIS
# ══════════════════════════════════════════════════════════════

def _variance_sheet(ws, df: pd.DataFrame, title: str, flagged_only: bool, unit_label: str):
    ws.sheet_view.showGridLines = False
    N = 7
    headers = ["Account Name", "Category", f"CY ({unit_label})", f"PY ({unit_label})",
               f"Change ({unit_label})", "Change (%)", "Variance Flag"]

    _h(ws, 1, N, f"{title} | Threshold based on data scale", fg=P["navy"], size=11, height=30)

    display_df = df.copy()
    if flagged_only:
        display_df = display_df[display_df["Variance Flag"].str.contains("🔴|🟡", na=False)]

    ws.row_dimensions[2].height = 14
    ws.merge_cells("A2:G2")
    note = ws["A2"]
    if flagged_only:
        note.value = f"  {len(display_df)} items require auditor attention — HIGH and MODERATE variance flags only"
    else:
        note.value = f"  All {len(display_df)} accounts shown | Sorted: flagged items first"
    note.font = Font(name="Calibri", size=9, italic=True, color=P["text_mid"])
    note.fill = PatternFill("solid", fgColor=P["grey_alt"])
    note.border = BORDER

    ws.row_dimensions[3].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=P["navy_mid"])
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, (_, row) in enumerate(display_df.iterrows(), start=4):
        flag = str(row.get("Variance Flag", ""))
        bg = _flag_bg(flag)
        if bg == P["white"]:
            bg = P["grey_alt"] if i % 2 == 0 else P["white"]
        ws.row_dimensions[i].height = 18

        row_vals = [
            row.get("Account Name", ""),
            row.get("Category", ""),
            row.get("CY Amount (₹)", 0),
            row.get("PY Amount (₹)", 0),
            row.get("Change (₹)", 0),
            row.get("Change (%)", None),
            flag,
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10, color=P["text_dark"])
            c.border = BORDER
            c.alignment = Alignment(
                horizontal="right" if col in [3, 4, 5, 6] else "left",
                vertical="center", indent=0 if col in [3,4,5,6] else 1
            )
            if col in [3, 4, 5] and isinstance(val, (int, float)):
                c.number_format = "#,##0.00"
            if col == 6 and isinstance(val, (int, float)) and val != 999.0:
                c.number_format = '+0.0;-0.0'

    _col_widths(ws, {1: 38, 2: 24, 3: 16, 4: 16, 5: 16, 6: 13, 7: 28})
    ws.freeze_panes = ws.cell(row=4, column=1)


# ══════════════════════════════════════════════════════════════
# SHEET 5 — GOING CONCERN
# ══════════════════════════════════════════════════════════════

def _going_concern_sheet(ws, gc: dict, standard: str):
    ws.title = "5. Going Concern"
    ws.sheet_view.showGridLines = False
    N = 5

    risk = gc["overall_risk"]
    score = gc["score"]
    risk_colors = {"CRITICAL": "991B1B", "HIGH": "C2410C", "MODERATE": "92400E", "LOW": "14532D"}
    risk_bg     = {"CRITICAL": "FEE2E2", "HIGH": "FEF3C7", "MODERATE": "FEF9C3", "LOW": "DCFCE7"}
    risk_icon   = {"CRITICAL": "🚨 CRITICAL", "HIGH": "🔴 HIGH", "MODERATE": "🟡 MODERATE", "LOW": "✅ LOW"}

    _h(ws, 1, N, f"GOING CONCERN ASSESSMENT — {standard} | ISA/SA 570",
       fg=P["navy"], size=12, height=32)

    # Overall verdict box
    ws.row_dimensions[2].height = 12
    ws.row_dimensions[3].height = 36
    ws.merge_cells("A3:E3")
    vc = ws["A3"]
    vc.value = f"OVERALL RISK LEVEL: {risk_icon.get(risk, risk)}   |   Risk Score: {score}/10"
    vc.fill  = PatternFill("solid", fgColor=risk_colors.get(risk, P["navy"]))
    vc.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = BORDER

    # Conclusion
    ws.row_dimensions[4].height = 56
    ws.merge_cells("A4:E4")
    cc = ws["A4"]
    cc.value = gc["conclusion"]
    cc.fill  = PatternFill("solid", fgColor=risk_bg.get(risk, P["grey_alt"]))
    cc.font  = Font(name="Calibri", size=10, color=P["text_dark"])
    cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cc.border = BORDER

    ws.row_dimensions[5].height = 10

    # Indicators header
    _h(ws, 6, N, "DETAILED INDICATORS", fg=P["navy_mid"], size=10, height=20)
    ws.row_dimensions[7].height = 20
    for col, h in enumerate(["ISA/SA 570 Indicator", "Status", "Finding", "", "Reference"], 1):
        c = ws.cell(row=7, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=P["slate"])
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for i, ind in enumerate(gc["indicators"], start=8):
        flag = str(ind.get("Status", ""))
        bg = _flag_bg(flag)
        ws.row_dimensions[i].height = 36

        vals = [
            ind.get("Indicator", ""),
            ind.get("Status", ""),
            ind.get("Finding", ""),
            "",
            ind.get("Reference", ""),
        ]
        # Merge col 3 + 4 for Finding (wider)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        for col, val in enumerate(vals, 1):
            if col == 4: continue
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10, color=P["text_dark"],
                           bold=(col == 1))
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)
            c.border = BORDER

    _col_widths(ws, {1: 38, 2: 22, 3: 50, 4: 4, 5: 32})


# ══════════════════════════════════════════════════════════════
# SHEET 6 — BENFORD'S LAW
# ══════════════════════════════════════════════════════════════

def _benfords_sheet(ws, bf: dict):
    ws.title = "6. Benford's Law"
    ws.sheet_view.showGridLines = False
    N = 7

    _h(ws, 1, N, "BENFORD'S LAW ANALYSIS — ISA/SA 240 Fraud Risk Indicator",
       fg=P["navy"], size=12, height=32)

    if not bf.get("sufficient", False):
        ws.row_dimensions[2].height = 60
        ws.merge_cells("A2:G2")
        c = ws["A2"]
        c.value = (
            f"⚠️ INSUFFICIENT DATA\n\n{bf.get('message', '')}\n\n"
            "To use Benford's Law analysis, upload a separate file of transaction-level amounts "
            "(journal entries, invoices, or ledger transactions) via the 'Benford's Law' tab in the app.\n\n"
            "Benford's Law requires individual transaction amounts — not aggregated trial balance totals."
        )
        c.fill = PatternFill("solid", fgColor=P["yellow_bg"])
        c.font = Font(name="Calibri", size=11, color=P["text_dark"])
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = BORDER
        return

    n = bf["n"]
    flag = bf.get("risk_flag", "")
    interp = bf.get("interpretation", "")

    # Result summary
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A2:G2")
    rc = ws["A2"]
    rc.value = f"{flag}   |   Dataset: {bf.get('label', 'Transactions')} ({n:,} amounts)"
    rc.fill  = PatternFill("solid", fgColor=_flag_bg(flag))
    rc.font  = Font(name="Calibri", size=12, bold=True, color=P["text_dark"])
    rc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rc.border = BORDER

    ws.row_dimensions[3].height = 54
    ws.merge_cells("A3:G3")
    ic = ws["A3"]
    ic.value = interp
    ic.fill  = PatternFill("solid", fgColor=P["grey_alt"])
    ic.font  = Font(name="Calibri", size=10, color=P["text_dark"])
    ic.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ic.border = BORDER

    ws.row_dimensions[4].height = 10

    # Data table
    _h(ws, 5, N, "DIGIT-BY-DIGIT ANALYSIS", fg=P["navy_mid"], size=10, height=20)
    headers = ["First Digit", "Observed Count", "Observed (%)", "Expected (%) — Benford",
               "Deviation (pp)", "Deviation (%)", "Flag"]
    ws.row_dimensions[6].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=P["slate"])
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(bf["digit_data"], start=7):
        dev = abs(row["Deviation (pp)"])
        if dev > 5:
            row_flag, bg = "🔴 Investigate", P["red_bg"]
        elif dev > 2:
            row_flag, bg = "🟡 Review", P["yellow_bg"]
        else:
            row_flag, bg = "🟢 OK", P["green_bg"]

        ws.row_dimensions[i].height = 18
        for col, val in enumerate([
            row["Digit"], row["Observed Count"], row["Observed (%)"],
            row["Expected (%)"], row["Deviation (pp)"], row["Deviation (%)"], row_flag
        ], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10, color=P["text_dark"])
            c.alignment = Alignment(horizontal="center" if col in [1,2] else "right",
                                    vertical="center")
            c.border = BORDER
            if col in [3, 4]:
                c.number_format = "0.00"
            if col == 5:
                c.number_format = "+0.00;-0.00"

    # Chi-square summary
    ws.row_dimensions[17].height = 10
    chi_row = 18
    _h(ws, chi_row, N, f"CHI-SQUARE STATISTIC = {bf.get('chi_square', 0):.2f}  "
       f"  |  Critical values: p<0.10 → 13.36  |  p<0.05 → 15.51  |  p<0.01 → 20.09  (df=8)",
       fg=P["slate"], size=9, height=20)

    _col_widths(ws, {1: 14, 2: 16, 3: 15, 4: 25, 5: 16, 6: 16, 7: 20})


# ══════════════════════════════════════════════════════════════
# SHEET 7 — INSTRUCTIONS
# ══════════════════════════════════════════════════════════════

def _instructions_sheet(ws, standard: str, unit_label: str):
    ws.title = "7. Instructions"
    ws.sheet_view.showGridLines = False

    _h(ws, 1, 6, "HOW TO USE THIS TOOL — Audit AP Tool v2.0", fg=P["navy"], size=12, height=28)

    content = [
        (True,  "QUICK START", None),
        (False, "1. Open the Streamlit app: run   streamlit run app_v2.py   in terminal", None),
        (False, "2. Enter client name, period, and select Ind AS or IFRS in the sidebar", None),
        (False, "3. Upload your trial balance Excel file (format shown below)", None),
        (False, "4. (Optional) Upload transaction amounts file for Benford's Law analysis", None),
        (False, "5. Click 'Generate Workpaper' — download formatted Excel file", None),
        (True,  "INPUT FORMAT — TRIAL BALANCE", None),
        (False, "Required columns (exact names):", None),
        (False, "  • Account Name   — account name as it appears in Tally", None),
        (False, "  • CY Amount (₹)  — current year balance (any unit: ₹, Lakhs, Crores)", None),
        (False, "  • PY Amount (₹)  — prior year comparative balance", None),
        (False, "Optional column:", None),
        (False, "  • Category       — auto-detected if blank (see valid values below)", None),
        (False, "IMPORTANT: Do NOT include subtotal rows (Total, Grand Total, etc.) — the tool skips them automatically", None),
        (True,  "VALID CATEGORY VALUES", None),
        (False, "  current_assets | fixed_assets | current_liabilities | long_term_liabilities", None),
        (False, "  equity | revenue | cogs | expenses | interest_expense", None),
        (True,  "BENFORD'S LAW INPUT FORMAT", None),
        (False, "Upload a separate Excel file with one column named 'Amount'", None),
        (False, "Each row = one transaction amount (journal entries, invoices, payments, etc.)", None),
        (False, "Needs at least 20 amounts. Works best with 500+ transactions.", None),
        (True,  "STANDARDS REFERENCE", None),
        (False, f"  {standard}:  Analytical Procedures → ISA/SA 520", None),
        (False, "  Going Concern         → ISA/SA 570 (6 indicators assessed)", None),
        (False, "  Fraud Risk            → ISA/SA 240 (Benford's Law)", None),
        (False, "  Risk Assessment       → ISA/SA 315", None),
        (False, "  Revenue Recognition   → IFRS 15 / Ind AS 115", None),
        (False, "  Impairment            → IAS 36 / Ind AS 36", None),
        (False, "  Provisions            → IAS 37 / Ind AS 37", None),
        (True,  "UNIT DETECTION", None),
        (False, f"  This workpaper detected amounts in: {unit_label}", None),
        (False, "  The tool auto-scales the materiality threshold based on detected unit scale.", None),
        (False, "  If incorrect, manually override in the sidebar before regenerating.", None),
        (True,  "INTERVIEW TALKING POINTS", None),
        (False, '"I built a tool that automates ISA 520 analytical procedures with IFRS and Ind AS support."', None),
        (False, '"It computes 15 financial ratios, flags material variances, and runs an ISA 570 going concern assessment."', None),
        (False, '"I also implemented Benford\'s Law (ISA 240) fraud detection on journal entries."', None),
        (False, '"The output is a formatted 7-sheet Excel workpaper — the kind filed in a Big 4 audit folder."', None),
        (False, "Built by Purva Doshi | Audit AP Tool v2.0 | ACCA Finalist", None),
    ]

    row = 2
    for bold, text, _ in content:
        ws.row_dimensions[row].height = 18 if bold else 15
        if bold:
            _h(ws, row, 6, f"  {text}", fg=P["divider"], font_color=P["text_dark"],
               size=10, bold=True, height=20)
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(name="Calibri", size=10, color=P["text_mid"])
            c.fill = PatternFill("solid", fgColor=P["white"])
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            c.border = BORDER
        row += 1

    _col_widths(ws, {1: 90})


# ══════════════════════════════════════════════════════════════
# MASTER REPORT BUILDER
# ══════════════════════════════════════════════════════════════

def generate_report(
    df_variance: pd.DataFrame,
    ratios: list,
    summary: dict,
    gc: dict,
    bf: dict,
    output_path: str,
    client_name: str = "Client",
    period: str = "FY 2024-25",
    standard: str = "Ind AS",
    unit_label: str = "₹",
) -> str:
    """Build and save the complete 7-sheet Excel workpaper."""
    _ensure_openpyxl()
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws1 = wb.create_sheet("1. Cover")
    ws2 = wb.create_sheet("2. Ratio Analysis")
    ws3 = wb.create_sheet("3. Variance")
    ws4 = wb.create_sheet("4. Flagged Items")
    ws5 = wb.create_sheet("5. Going Concern")
    ws6 = wb.create_sheet("6. Benford's Law")
    ws7 = wb.create_sheet("7. Instructions")

    _cover_sheet(ws1, summary, client_name, period, unit_label, standard)
    _ratio_sheet(ws2, ratios, standard, unit_label)
    _variance_sheet(ws3, df_variance, "3. LINE-ITEM VARIANCE — All Accounts", False, unit_label)
    _variance_sheet(ws4, df_variance, "4. FLAGGED ITEMS — Auditor Focus", True, unit_label)
    _going_concern_sheet(ws5, gc, standard)
    _benfords_sheet(ws6, bf)
    _instructions_sheet(ws7, standard, unit_label)

    wb.save(output_path)
    return output_path
